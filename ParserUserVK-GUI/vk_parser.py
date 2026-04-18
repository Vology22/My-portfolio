import vk_api
import time
import json
from urllib.parse import urlparse
import openpyxl
from datetime import datetime
import os



def extract_vk_name(url):
    parsed_url = urlparse(url)
    path = parsed_url.path.strip('/')
    
    # Разделяем путь и возможные параметры пути (если есть)
    name = path.split('/')[0]
    
    return name
def get_all_group_members(group_id, city, vk):
    
    # Сначала получаем общее количество участников
    group_info = vk.groups.getMembers(group_id=group_id, count=0)
    total_members = group_info["count"]
    print(f"Всего участников: {total_members}")

    members = []
    offset = 0
    batch_size = 1000  # Максимально допустимое значение
    sort_members = []
    while offset < total_members:
        try:
            response = vk.groups.getMembers(
                group_id=group_id,
                offset=offset,
                count=batch_size,
                fields="first_name,last_name,city,domain" # Опциональные поля
            )

            members.extend(response["items"])
            offset += batch_size
            print(f"Загружено: {len(members)} / {total_members}")
            
            # Задержка, чтобы не превысить лимиты API (3 запроса/сек)
            time.sleep(0.2)  
             
        except vk_api.exceptions.ApiError as e:
            print(f"Ошибка: {e}")
            break
    if not (city == ["-"]):
        for user in members:
            if "city" in user.keys():
                for cc in city:
                    if user["city"]["id"] == int(cc):
                        sort_members.append(user)
                        break
    else:
        for user in members:
            sort_members.append(user)
    
    return sort_members
def save_result_to_excel(base):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parsing Results VK"
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="domain")
    ws.cell(row=1, column=3, value="city")
    ws.cell(row=1, column=4, value="name")
    for user in range(len(base)):
        id_user = base[user]["id"]
        domain_user = "https://vk.com/" + base[user]["domain"]
        try:
            city_user = base[user]["city"]["title"]
        except:
            city_user = "Close"
        name_user = base[user]['first_name'] + " " + base[user]['last_name']
        ws.cell(row=user+2, column=1, value=id_user)
        ws.cell(row=user+2, column=2, value=domain_user)
        ws.cell(row=user+2, column=3, value=city_user)
        ws.cell(row=user+2, column=4, value=name_user)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width
    os.makedirs("./output", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"./output/parsed_results_{timestamp}.xlsx"
    wb.save(filename)
    print(f"Результаты сохранены в {filename}")
    return filename
#'4ea20b134ea20b134ea20b13b74d9266c544ea24ea20b1326adf29140f17febf3101aa7'
def fucking_parse_vk(city, groups, token):
    vk_session = vk_api.VkApi(token=token)  # Если нужен доступ к методам API
    vk = vk_session.get_api()
    cities = []
    if city == ["-"]:
        cities = ["-"]
        print("-")
    else:
        for i in city:
            strana = vk.database.getCities(need_all=0, q=i, count=1)["items"][0]["id"]
            cities.append(strana)
    print(cities)
    #{'id': 1023836928, 'domain': 'id1023836928', 'city': {'id': 1, 'title': 'Moscow'}, 'first_name': 'Oleg', 'last_name': 'Vetrov', 'can_access_closed': True, 'is_closed': False}
    base = []
    for i in groups:
        a = get_all_group_members(extract_vk_name(i), cities, vk)
        for i in a:
            base.append(i)
    return save_result_to_excel(base)
if __name__ == "__main__":
   fucking_parse_vk("-", ["https://vk.com/sch28hodec?from=groups"], '4ea20b134ea20b134ea20b13b74d9266c544ea24ea20b1326adf29140f17febf3101aa7')



