########################################################## Parser by Vology #############################################################################
# Привет, мой дорогой друг. Если ты читаешь этот комментарий, то скорее всего ты работаешь над парсером, который писал я.
# Уверен, во время работы над ним у тебя появятся вопросы... Много вопросов. 
# Если тебе потребуется помощь, пиши сюда TG @onevology
# Удачи! <3

from bs4 import BeautifulSoup as bs
import time
import os
import requests
import openpyxl
import random
from time import sleep

headers = {
    'Accept-Encoding': 'gzip, deflate, sdch',
    'Accept-Language': 'en-US,en;q=0.8',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
}
def joker(text):
      linee = split(text)
      while linee:
            sleep(0.1)
            print(linee[0], end='')
            linee.remove(linee[0])
def split(s):
    return [char for char in s]
joker("Parser by @onevology (TG)")
#првиет как дела я люблю когда волосатый мужики обмазываются маслом
print("\n\n")
#  Название или ФИО, сайт, соцсети, почта и телефон
def parse(path, st):   # res - количество страниц      city:   0 - Москва    1 - Питер      arhdiz:    0 - Архитекторы    1 - Дизайнеры
    arh_houz = []
    osn_url = ''
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active   
    cell_obj = sheet_obj.cell(row = 1, column = 1) #cell_obj.value
    for i in range(sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row = i + 2, column = st)
        if cell_obj.value == "":
            break
        print(cell_obj.value)
        if cell_obj.value is None:
            continue
        arh_houz.append(cell_obj.value)
    
    print(f"Найдено профилей: "+str(len(arh_houz)))
    print("Сейчас начнется сбор html-копий")
    pages = []
    chet = 0
    for i in arh_houz:
        print(f"Сбор html-копий профилей {chet}/{str(len(arh_houz))}", end="\r")
        r = requests.get(i, headers=headers)
        pages.append([r.text, i])
        chet += 1
    print(f"Сбор html-копий профилей {chet}/{str(len(arh_houz))}")
    
    #### парсинг основных данных ####
    print("\n")
    print("Сейчас начнется сбор данных...")
    data_base = []
    chet = 0
    for i in pages:
        soup = bs(i[0], "html.parser")
        #try:
        main_block = soup.find_all('div', class_='sc-183mtny-0 sc-1wm9uar-0 lnQdrs bHBpBq hui-grid')[0].find_all('p')
        name = '-'
        name = main_block[0].text # name 
        number = '-'
        number = main_block[1].text # number
        if number == "" or number == " " or number is None:
            number = '-'
        website = '-'
        website = main_block[2].text # site
        try:
            r = requests.get(website, headers=headers)
            if not r.status_code == 200:
                    website = "-"
        except:
            try:
                r = requests.get("https://" + website, headers=headers)
                if not r.status_code == 200:
                    website = "-"
                    #<div class="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 hTSLyQ hNCBYN jlkdvL hui-cell">
            except:
                website = "-"
        main_block = soup.find_all('div', class_='sc-183mtny-0 sc-1wm9uar-0 lnQdrs bHBpBq hui-grid')[0].find_all('div', class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 hTSLyQ hNCBYN jlkdvL hui-cell")
        if main_block[len(main_block)-1].find("h3").text == "Socials":
            soc = main_block[len(main_block)-1].find("a")["href"]
        else:
            soc = "-"

        data_base.append([name, number, website, i[1], soc]) # имя, номер, сайт, url
        chet += 1
        #except:
            #continue
        print(f"Парсинг основных данных {chet}/{str(len(pages))}", end="\r")
    print(f"Парсинг основных данных {chet}/{str(len(pages))}")
    print("Last Process")
    print("Формируем Excel-таблицу")
    wb = openpyxl.Workbook()
    list1 = wb.active
    list1.cell(row=1, column=1, value="Название")
    list1.cell(row=1, column=2, value="Номер")
    list1.cell(row=1, column=3, value="URL-адрес")
    list1.cell(row=1, column=4, value="Соц сети (inst)")
    list1.cell(row=1, column=5, value="Сайт")

    for i in range(len(data_base)):
        if i == 0:
            continue
        list1.cell(row=i+1, column=1, value=data_base[i][0])
        list1.cell(row=i+1, column=2, value=data_base[i][1])
        list1.cell(row=i+1, column=3, value=data_base[i][3])
        list1.cell(row=i+1, column=4, value=data_base[i][4])# соц сети
        list1.cell(row=i+1, column=5, value=data_base[i][2])

    name = str(random.randint(1000000, 9999999))
    wb.save(f"db_{name}_houzz.xlsx")
    print(f"Ваша база данных успешно собрана! \nНазвание: db_{name}_houzz.xlsx")
print("\n")
path_db = input("Введите название базы данных (она должна быть в одном каталоге со скриптом): ")
print("\n")
st = int(input("Введите номер столбика, где хранятся URL(например 2): "))
print("Приветствую, это парсер сайта www.houzz.ru")
parse(path_db, st)
print("\nНажмите Enter, чтобы выйти")
input()

