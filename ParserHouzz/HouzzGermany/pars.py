########################################################## Parser by Vology #############################################################################
# Привет, мой дорогой друг. Если ты читаешь этот комментарий, то скорее всего ты работаешь над парсером, который писал я.
# Уверен, во время работы над ним у тебя появятся вопросы... Много вопросов. 
# Если тебе потребуется помощь, пиши сюда TG @onevology
# Удачи! <3
import os
from msedge.selenium_tools import Edge, EdgeOptions
from bs4 import BeautifulSoup as bs
import time
import os
import requests
import openpyxl
import random
from time import sleep

arh_citis = [["Aachen", "https://www.houzz.de/professionals/architekten/c/Aachen--Nordrhein_Westfalen"], ["Augsburg", "https://www.houzz.de/professionals/architekten/c/Augsburg--Bayern"], ["Basel", "https://www.houzz.de/professionals/architekten/c/Basel--Kanton-Basel_Stadt--Schweiz"], ["Berlin", "https://www.houzz.de/professionals/architekten/c/Berlin"], ["Bielefeld", "https://www.houzz.de/professionals/architekten/c/Bielefeld--Nordrhein_Westfalen"], ["Bonn", "https://www.houzz.de/professionals/architekten/c/Bonn--Nordrhein_Westfalen"], ["Braunschweig", "https://www.houzz.de/professionals/architekten/c/Braunschweig--Niedersachsen"], ["Bremen", "https://www.houzz.de/professionals/architekten/c/Bremen"], ["Chemnitz", "https://www.houzz.de/professionals/architekten/c/Chemnitz--Sachsen"], ["Darmstadt", "https://www.houzz.de/professionals/architekten/c/Darmstadt--Hessen"], ["Dortmund", "https://www.houzz.de/professionals/architekten/c/Dortmund--Nordrhein_Westfalen"], ["Dresden", "https://www.houzz.de/professionals/architekten/c/Dresden--Sachsen"], ["Duisburg", "https://www.houzz.de/professionals/architekten/c/Duisburg--Nordrhein_Westfalen"], ["Düsseldorf", "https://www.houzz.de/professionals/architekten/c/Düsseldorf--Nordrhein_Westfalen"], ["Erfurt", "https://www.houzz.de/professionals/architekten/c/Erfurt--Thüringen"], ["Essen", "https://www.houzz.de/professionals/architekten/c/Essen--Nordrhein_Westfalen"], ["Frankfurt am Main", "https://www.houzz.de/professionals/architekten/c/Frankfurt-am-Main--Hessen"], ["Freiburg-im-Breisgau", "https://www.houzz.de/professionals/architekten/c/Freiburg-im-Breisgau--Baden_Württemberg"], ["Fürth", "https://www.houzz.de/professionals/architekten/c/Fürth--Bayern"], ["Gelsenkirchen", "https://www.houzz.de/professionals/architekten/c/Gelsenkirchen--Nordrhein_Westfalen"], ["Graz", "https://www.houzz.de/professionals/architekten/c/Graz--Steiermark--Österreich"], ["Hagen", "https://www.houzz.de/professionals/architekten/c/Hagen--Nordrhein_Westfalen"], ["Halle", "https://www.houzz.de/professionals/architekten/c/Halle-(Saale)--Sachsen_Anhalt"], ["Hamburg", "https://www.houzz.de/professionals/architekten/c/Hamburg"], ["Hamm", "https://www.houzz.de/professionals/architekten/c/Hamm--Nordrhein_Westfalen"], ["Hannover", "https://www.houzz.de/professionals/architekten/c/Hannover--Niedersachsen"], ["Heidelberg", "https://www.houzz.de/professionals/architekten/c/Heidelberg--Baden_Württemberg"], ["Herne", "https://www.houzz.de/professionals/architekten/c/Herne--Nordrhein_Westfalen"], ["Ingolstadt", "https://www.houzz.de/professionals/architekten/c/Ingolstadt--Bayern"], ["Karlsruhe", "https://www.houzz.de/professionals/architekten/c/Karlsruhe--Baden_Württemberg"], ["Kassel", "https://www.houzz.de/professionals/architekten/c/Kassel--Hessen"], ["Kiel", "https://www.houzz.de/professionals/architekten/c/Kiel--Schleswig_Holstein"], ["Krefeld", "https://www.houzz.de/professionals/architekten/c/Krefeld--Nordrhein_Westfalen"], ["Köln", "https://www.houzz.de/professionals/architekten/c/Köln--Nordrhein_Westfalen"], ["Leipzig", "https://www.houzz.de/professionals/architekten/c/Leipzig--Sachsen"], ["Leverkusen", "https://www.houzz.de/professionals/architekten/c/Leverkusen--Nordrhein_Westfalen"], ["Ludwigshafen-am-Rhein", "https://www.houzz.de/professionals/architekten/c/Ludwigshafen-am-Rhein--Rheinland_Pfalz"], ["Lübeck", "https://www.houzz.de/professionals/architekten/c/Lübeck--Schleswig_Holstein"], ["Magdeburg", "https://www.houzz.de/professionals/architekten/c/Magdeburg--Sachsen_Anhalt"], ["Mainz", "https://www.houzz.de/professionals/architekten/c/Mainz--Rheinland_Pfalz"], ["Mannheim", "https://www.houzz.de/professionals/architekten/c/Mannheim--Baden_Württemberg"], ["Mönchen­gladbach", "https://www.houzz.de/professionals/architekten/c/Mönchengladbach--Nordrhein_Westfalen"], ["Mülheim-an-der-Ruhr", "https://www.houzz.de/professionals/architekten/c/Mülheim-(Ruhr)--Nordrhein_Westfalen"], ["München", "https://www.houzz.de/professionals/architekten/c/München--Bayern"], ["Münster", "https://www.houzz.de/professionals/architekten/c/Münster--Nordrhein_Westfalen"], ["Neuss", "https://www.houzz.de/professionals/architekten/c/Neuss--Nordrhein_Westfalen"], ["Nürnberg", "https://www.houzz.de/professionals/architekten/c/Nürnberg--Bayern"], ["Oberhausen", "https://www.houzz.de/professionals/architekten/c/Oberhausen--Nordrhein_Westfalen"], ["Offenbach-am-Main", "https://www.houzz.de/professionals/architekten/c/Offenbach-am-Main--Hessen"], ["Oldenburg", "https://www.houzz.de/professionals/architekten/c/Oldenburg-(Oldb)--Niedersachsen"], ["Osnabrück", "https://www.houzz.de/professionals/architekten/c/Osnabrück--Niedersachsen"], ["Paderborn", "https://www.houzz.de/professionals/architekten/c/Paderborn--Nordrhein_Westfalen"], ["Potsdam", "https://www.houzz.de/professionals/architekten/c/Potsdam--Brandenburg"], ["Regensburg", "https://www.houzz.de/professionals/architekten/c/Regensburg--Bayern"], ["Rostock", "https://www.houzz.de/professionals/architekten/c/Rostock--Mecklenburg_Vorpommern"], ["Saarbrücken", "https://www.houzz.de/professionals/architekten/c/Saarbrücken--Saarland"], ["Solingen", "https://www.houzz.de/professionals/architekten/c/Solingen--Nordrhein_Westfalen"], ["Stuttgart", "https://www.houzz.de/professionals/architekten/c/Stuttgart--Baden_Württemberg"], ["Wien", "https://www.houzz.de/professionals/architekten/c/Wien--Wien--Österreich"], ["Wiesbaden", "https://www.houzz.de/professionals/architekten/c/Wiesbaden--Hessen"], ["Wuppertal", "https://www.houzz.de/professionals/architekten/c/Wuppertal--Nordrhein_Westfalen"], ["Würzburg", "https://www.houzz.de/professionals/architekten/c/Würzburg--Bayern"], ["Zürich", "https://www.houzz.de/professionals/architekten/c/Zürich--Kanton-Zürich--Schweiz"]]
diz_citis = [["Aachen", "https://www.houzz.de/professionals/innenarchitekten/c/Aachen--Nordrhein_Westfalen"], ["Städteregion Aachen", "https://www.houzz.de/professionals/innenarchitekten/c/Städteregion-Aachen--Nordrhein_Westfalen"], ["Stolberg", "https://www.houzz.de/professionals/innenarchitekten/c/Stolberg--Nordrhein_Westfalen"], ["Eschweiler", "https://www.houzz.de/professionals/innenarchitekten/c/Eschweiler--Nordrhein_Westfalen"], ["Herzogenrath", "https://www.houzz.de/professionals/innenarchitekten/c/Herzogenrath--Nordrhein_Westfalen"], ["Alsdorf", "https://www.houzz.de/professionals/innenarchitekten/c/Alsdorf--Nordrhein_Westfalen"], ["Würselen", "https://www.houzz.de/professionals/innenarchitekten/c/Würselen--Nordrhein_Westfalen"], ["Jülich", "https://www.houzz.de/professionals/innenarchitekten/c/Jülich--Nordrhein_Westfalen"], ["Geilenkirchen", "https://www.houzz.de/professionals/innenarchitekten/c/Geilenkirchen--Nordrhein_Westfalen"], ["Baesweiler", "https://www.houzz.de/professionals/innenarchitekten/c/Baesweiler--Nordrhein_Westfalen"], ["Linnich", "https://www.houzz.de/professionals/innenarchitekten/c/Linnich--Nordrhein_Westfalen"], ["Siersdorf", "https://www.houzz.de/professionals/innenarchitekten/c/Siersdorf--Nordrhein_Westfalen"], ["Neu-Pattern", "https://www.houzz.de/professionals/innenarchitekten/c/Neu_Pattern--Nordrhein_Westfalen"], ["Niedermerz", "https://www.houzz.de/professionals/innenarchitekten/c/Niedermerz--Nordrhein_Westfalen"], ["Freialdenhoven", "https://www.houzz.de/professionals/innenarchitekten/c/Freialdenhoven--Nordrhein_Westfalen"], ["Breinigerberg", "https://www.houzz.de/professionals/innenarchitekten/c/Breinigerberg--Nordrhein_Westfalen"], ["Rurberg", "https://www.houzz.de/professionals/innenarchitekten/c/Rurberg--Nordrhein_Westfalen"], ["Dürboslar", "https://www.houzz.de/professionals/innenarchitekten/c/Dürboslar--Nordrhein_Westfalen"], ["Einruhr", "https://www.houzz.de/professionals/innenarchitekten/c/Einruhr--Nordrhein_Westfalen"], ["Puffendorf", "https://www.houzz.de/professionals/innenarchitekten/c/Puffendorf--Nordrhein_Westfalen"], ["Engelsdorf", "https://www.houzz.de/professionals/innenarchitekten/c/Engelsdorf--Nordrhein_Westfalen"], ["Brandenberg", "https://www.houzz.de/professionals/innenarchitekten/c/Brandenberg--Nordrhein_Westfalen"], ["Friesenrath", "https://www.houzz.de/professionals/innenarchitekten/c/Friesenrath--Nordrhein_Westfalen"], ["Walheim", "https://www.houzz.de/professionals/innenarchitekten/c/Walheim--Nordrhein_Westfalen"], ["Mulartshütte", "https://www.houzz.de/professionals/innenarchitekten/c/Mulartshütte--Nordrhein_Westfalen"], ["Marienthal", "https://www.houzz.de/professionals/innenarchitekten/c/Marienthal--Nordrhein_Westfalen"], ["Lichtenbusch", "https://www.houzz.de/professionals/innenarchitekten/c/Lichtenbusch--Nordrhein_Westfalen"], ["Nütheim", "https://www.houzz.de/professionals/innenarchitekten/c/Nütheim--Nordrhein_Westfalen"], ["Zweifall", "https://www.houzz.de/professionals/innenarchitekten/c/Zweifall--Nordrhein_Westfalen"], ["Schleckheim", "https://www.houzz.de/professionals/innenarchitekten/c/Schleckheim--Nordrhein_Westfalen"]]
headers = {
    'Accept-Encoding': 'gzip, deflate, sdch',
    'Accept-Language': 'en-US,en;q=0.8',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
}
#try:
edge_options = EdgeOptions()
edge_options.use_chromium = True 
edge_options.add_argument("--start-maximized")
edge_options.add_argument("--disable-dev-shm-usage")
edge_options.add_argument("--no-sandbox") 
edge_options.add_argument("--remote-debugging-port=9222") 
edge_options.add_argument('--log-level=3')
with open("config.txt", "r") as f:
    path = f.readlines()[0].replace("\n", "")
    edge_options.add_argument(f"user-data-dir={path}")
edge_options.add_argument(f"profile-directory=Profile 1")
#edge_options.add_argument("headless")
#edge_options.add_argument("disable-gpu")
def joker(text):
    linee = split(text)
    while linee:
            sleep(0.1)
            print(linee[0], end='')
            linee.remove(linee[0])
def split(s):
    return [char for char in s]
joker("Parser by @onevology (TG)")
print("\n\n")
#  Название или ФИО, сайт, соцсети, почта и телефон
def parse(res, city, arhdiz):   # res - количество страниц      city:   0 - Москва    1 - Питер      arhdiz:    0 - Архитекторы    1 - Дизайнеры
    arh_houz = []
    osn_url = ''
    driver = Edge(options=edge_options, executable_path = os.getcwd() + "\\msedgedriver.exe")
    for i in range(res):
        if city == "test":
            if arhdiz == 0:
                if i == 0:
                    osn_url = "https://www.houzz.ru/professionals/Arkhitektory#"
                    URL_TEMPLATE = osn_url
                else:
                    page_number = i * 15
                    URL_TEMPLATE = osn_url + "/p/" + str(page_number)
            elif arhdiz == 1:
                if i == 0:
                    osn_url = "https://www.houzz.ru/professionals/Arkhitektory#"
                    URL_TEMPLATE = osn_url
                    URL_TEMPLATE = osn_url
                else:
                    page_number = i * 15
                    URL_TEMPLATE = osn_url + "/p/" + str(page_number)
        elif arhdiz == 0:
            if i == 0:
                osn_url = arh_citis[city][1]
                URL_TEMPLATE = osn_url
            else:
                page_number = i * 15
                URL_TEMPLATE = osn_url + "/p/" + str(page_number)
        elif arhdiz == 1:
            if i == 0:
                osn_url = diz_citis[city][1]
                URL_TEMPLATE = osn_url
                URL_TEMPLATE = osn_url
            else:
                page_number = i * 15
                URL_TEMPLATE = osn_url + "/p/" + str(page_number)
        print(URL_TEMPLATE)
        driver.get(URL_TEMPLATE)
        #r = requests.get(URL_TEMPLATE, headers=headers)
        time.sleep(1)
        #print(r.text)
        
        with open("test.html", "r+", encoding="utf-8") as f:
           f.write(driver.page_source)
        soup = bs(driver.page_source, "html.parser")
        
        #<li class="ProSearchResultsV2__StyledListItem-aeq0am-0 hz-pro-search-results__item ">
        #<li class="ProSearchResultsV2__StyledListItem-aeq0am-0 hz-pro-search-results__item ">
        #<li class="ProSearchResultsV2__StyledListItem-aeq0am-0 hz-pro-search-results__item ">
        #print(len(soup.find_all('div', class_='pro-results')[0].find_all('li', class_='ProSearchResultsV2__StyledListItem-aeq0am-0 hz-pro-search-results__item ')))
        #<div class="sc-183mtny-0 sc-dpAhYB dbBdzY bexXIR separate-box">
        #<ul class="hz-pro-search-results mb0">
        #print(len(soup.find_all('ul', class_='hz-pro-search-results mb0')[0]))
        arh = soup.find_all('ul', class_='hz-pro-search-results mb0')[0].find_all('li')
        # <a class="sc-62xgu6-0 cYgEZl sc-mwxddt-0 iIGnLB hui-link sc-bdvvtL ksqaeH hz-pro-ctl" href="https://www.houzz.ru/pro/de-club/de-club" font-size="inherit,inherit">
        for i in arh:
                #print(len(i.find_all("a")))

                arh_houz.append(i.find_all("a")[0]["href"])
                print(i.find_all("a")[0]["href"])
    
    print(f"Найдено профилей с {res} страниц: "+str(len(arh_houz)))
    print("Сейчас начнется сбор html-копий")
    pages = []
    chet = 0
    
    driver.implicitly_wait(10) # Ждет до 10 секунд
    for i in arh_houz:
        print(f"Сбор html-копий профилей {chet}/{str(len(arh_houz))}    ===========   {i}")#, end="\r") 
        driver.get(i)
        pages.append([driver.page_source, i])
        chet += 1
    print(f"Сбор html-копий профилей {chet}/{str(len(arh_houz))}")
    driver.close()
    
    #### парсинг основных данных ####
    print("\n")
    print("Сейчас начнется сбор данных...")
    data_base = []
    chet = 0
    for i in pages:
        soup = bs(i[0], "html.parser")
        try:
            main_block = soup.find_all('div', class_='sc-183mtny-0 sc-1wm9uar-0 lnQdrs bHBpBq hui-grid')[0].find_all('p')
            name = '-'
            name = main_block[0].text # name 
            number = '-'
            number = main_block[1].text # number
            if number == "" or number == " " or number is None:
                number = '-'
            website = '-'
            website = main_block[2].text # site
            try:
                a = int(website[0]) + int(website[1]) 
                website = "-"
            except Exception as e:
                a = 0
                
            adress = '-'
            adress = main_block[3].text # adress
            try:
                a = int(adress[0]) + int(adress[2]) 
                adress = "-"
            except Exception as e:
                a = 0
            if adress == "" or adress == " " or adress is None:
                adress = '-'
            data_base.append([name, number, website, adress, i[1]]) # имя, номер, сайт, адрес, url
            chet += 1
        except:
            continue
        print(f"Парсинг основных данных {chet}/{str(len(pages))}", end="\r")
    print(f"Парсинг основных данных {chet}/{str(len(pages))}")
    print("Last Process")
    print("Формируем Excel-таблицу")
    wb = openpyxl.Workbook()
    list1 = wb.active
    list1.cell(row=1, column=1, value="Название")
    list1.cell(row=1, column=2, value="Номер")
    list1.cell(row=1, column=3, value="Сайт")
    list1.cell(row=1, column=4, value="Адрес")
    list1.cell(row=1, column=5, value="URL-адрес")
    list1.cell(row=1, column=7, value="city")
    try:
        if arhdiz == 0:
            prof = "Architects"
            oh_city = arh_citis[city][0]
        elif arhdiz == 1:
            oh_city = diz_citis[city][0]
            prof = "Designers"
    except:
        prof = "Arh"
        oh_city = 'test'
    list1.cell(row=2, column=7, value=oh_city)
    list1.cell(row=1, column=8, value="prof")
    list1.cell(row=2, column=8, value=prof)
    for i in range(len(data_base)):
        if i == 0:
            continue
        list1.cell(row=i+1, column=1, value=data_base[i][0])
        list1.cell(row=i+1, column=2, value=data_base[i][1])
        list1.cell(row=i+1, column=3, value=data_base[i][2])
        list1.cell(row=i+1, column=4, value=data_base[i][3])
        list1.cell(row=i+1, column=5, value=data_base[i][4])
        list1.cell(row=i+1, column=5, value=data_base[i][4])

        list1.cell(row=i+2, column=8, value=".")
        list1.cell(row=i+2, column=7, value=".")
    name = str(random.randint(1000000, 9999999))
    wb.save(f"db_{name}_{prof}_{oh_city}_houzz.xlsx")
    print(f"Ваша база данных успешно собрана! \nНазвание: db_{name}_{prof}_{oh_city}_houzz.xlsx")
print("\n")
print("Приветствую, это парсер сайта www.houzz.ru")
page_cite = int(input("Введите количество страниц (целое число, 1 = +- 14 страниц. Например 4 значит 56 профилей): "))
print("\n                                                                                              ", end="\r")
arh_cite = int(input("Направление (0 - Архитекторы   1 - Дизайнеры): "))
while not arh_cite == 1 and not arh_cite == 0:
    print("\n                                                                                 ", end="\r")
    arh_cite = int(input("Направление (0 - Архитекторы   1 - Дизайнеры): "))
if arh_cite == 0:
    for i in range(len(arh_citis)):
        print(f"{i}. {arh_citis[i][0]}")
elif arh_cite == 1:
    for i in range(len(diz_citis)):
        print(f"{i}. {diz_citis[i][0]}")
else:
    print("Вы выбрали неверное значение :(")
    exit()


city_cite = input("Введите индекс города: ")
print("\n                                                                                              ", end="\r")
if not city_cite == "test":
    city_cite = int(city_cite)


print("Начинаем парсинг :D")
parse(page_cite, city_cite, arh_cite)
print("\nНажмите Enter, чтобы выйти")
input()

#except Exception as e:
    #print(f"Произошла непредвиденная ошибка: {e}")
    #input("Нажмите Enter, чтобы выйти")