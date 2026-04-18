########################################################## Parser by Vology #############################################################################
# Привет, мой дорогой друг. Если ты читаешь этот комментарий, то скорее всего ты работаешь над парсером, который писал я.
# Уверен, во время работы над ним у тебя появятся вопросы... Много вопросов. 
# Если тебе потребуется помощь, пиши сюда TG @onevology
# Удачи! <3

from bs4 import BeautifulSoup as bs
import time
import os
import requests
import openpyxl
import random
from time import sleep
arh_citis = [["Москва", "https://www.houzz.ru/professionals/Arkhitektory/c/Москва--регион-Москва"], ["Питер", "https://www.houzz.ru/professionals/Arkhitektory/c/Санкт_Петербург--регион-Санкт_Петербург"], ["Новосибирск", "https://www.houzz.ru/professionals/Arkhitektory/c/Новосибирск--Новосибирская-область"], ["Екатеринбург", "https://www.houzz.ru/professionals/Arkhitektory/c/Екатеринбург--Свердловская-область"], ["Нижний Новгород", "https://www.houzz.ru/professionals/Arkhitektory/c/Нижний-Новгород--Нижегородская-область"], ["Казать", "https://www.houzz.ru/professionals/Arkhitektory/c/Казань--Татарстан"], ["Ижевск", "https://www.houzz.ru/professionals/Arkhitektory/c/Ижевск--Удмуртия"], ["Омск", "https://www.houzz.ru/professionals/Arkhitektory/c/Омск--Омская-область"], ["Самара", "https://www.houzz.ru/professionals/Arkhitektory/c/Самара--Самарская-область"], ["Ростов на Дону", "https://www.houzz.ru/professionals/Arkhitektory/c/Ростов_на_Дону--Ростовская-область"], ["Уфа", "https://www.houzz.ru/professionals/Arkhitektory/c/Уфа--Башкортостан"], ["Красноярск", "https://www.houzz.ru/professionals/Arkhitektory/c/Красноярск--Красноярский-край"], ["Воронеж", "https://www.houzz.ru/professionals/Arkhitektory/c/Воронеж--Воронежская-область"], ["Пермь", "https://www.houzz.ru/professionals/Arkhitektory/c/Пермь--Пермская-область"], ["Волгоград", "https://www.houzz.ru/professionals/Arkhitektory/c/Волгоград--Волгоградская-область"], ["Краснодар", "https://www.houzz.ru/professionals/Arkhitektory/c/Краснодар--Краснодарский-край"], ["Саратов", "https://www.houzz.ru/professionals/Arkhitektory/c/Саратов--Саратовская-область"], ["Тюмень", "https://www.houzz.ru/professionals/Arkhitektory/c/Тюмень--Тюменская-область"], ["Ижевск", "https://www.houzz.ru/professionals/Arkhitektory/c/Ижевск--Удмуртия"], ["Барнаул", "https://www.houzz.ru/professionals/Arkhitektory/c/Барнаул--Алтайский-край"], ["Иркутск", "https://www.houzz.ru/professionals/Arkhitektory/c/Иркутск--Иркутская-область"], ["Хабаровск", "https://www.houzz.ru/professionals/Arkhitektory/c/Хабаровск--Хабаровский-край"], ["Ярославль", "https://www.houzz.ru/professionals/Arkhitektory/c/Ярославль--Ярославская-область"], ["Владивосток", "https://www.houzz.ru/professionals/Arkhitektory/c/Владивосток--Приморский-край"], ["Махачкала", "https://www.houzz.ru/professionals/Arkhitektory/c/Махачкала--Дагестан"], ["Томск", "https://www.houzz.ru/professionals/Arkhitektory/c/Томск--Томская-область"], ["Рязань", "https://www.houzz.ru/professionals/Arkhitektory/c/Рязань--Рязанская-область"], ["Пенза", "https://www.houzz.ru/professionals/Arkhitektory/c/Пенза--Пензенская-область"], ["Киров", "https://www.houzz.ru/professionals/Arkhitektory/c/Киров--Кировская-область"], ["Чебоксары", "https://www.houzz.ru/professionals/Arkhitektory/c/Чебоксары--Чувашия"], ["Калининград", "https://www.houzz.ru/professionals/Arkhitektory/c/Калининград--Калининградская-область"], ["Тула", "https://www.houzz.ru/professionals/Arkhitektory/c/Тула--Тульская-область"], ["Сочи", "https://www.houzz.ru/professionals/Arkhitektory/c/Сочи--Краснодарский-край"], ["Ставрополь", "https://www.houzz.ru/professionals/Arkhitektory/c/Ставрополь--Ставрополье"], ["Тверь", "https://www.houzz.ru/professionals/Arkhitektory/c/Тверь--Тверская-область"]]
diz_citis = [["Москва", "https://www.houzz.ru/professionals/dizayn-interyera/c/Москва--регион-Москва"], ["Санкт-Петербург", "https://www.houzz.ru/professionals/dizayn-interyera/c/Санкт_Петербург--регион-Санкт_Петербург"], ["Новосибирск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Новосибирск--Новосибирская-область"], ["Екатеринбург", "https://www.houzz.ru/professionals/dizayn-interyera/c/Екатеринбург--Свердловская-область"], ["Нижний Новгород", "https://www.houzz.ru/professionals/dizayn-interyera/c/Нижний-Новгород--Нижегородская-область"], ["Казань", "https://www.houzz.ru/professionals/dizayn-interyera/c/Казань--Татарстан"], ["Челябинск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Ижевск--Удмуртия"], ["Омск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Омск--Омская-область"], ["Самара", "https://www.houzz.ru/professionals/dizayn-interyera/c/Самара--Самарская-область"], ["Ростов-на-Дону", "https://www.houzz.ru/professionals/dizayn-interyera/c/Ростов_на_Дону--Ростовская-область"], ["Уфа", "https://www.houzz.ru/professionals/dizayn-interyera/c/Уфа--Башкортостан"], ["Красноярск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Красноярск--Красноярский-край"], ["Воронеж", "https://www.houzz.ru/professionals/dizayn-interyera/c/Воронеж--Воронежская-область"], ["Пермь", "https://www.houzz.ru/professionals/dizayn-interyera/c/Пермь--Пермская-область"], ["Волгоград", "https://www.houzz.ru/professionals/dizayn-interyera/c/Волгоград--Волгоградская-область"], ["Краснодар", "https://www.houzz.ru/professionals/dizayn-interyera/c/Краснодар--Краснодарский-край"], ["Саратов", "https://www.houzz.ru/professionals/dizayn-interyera/c/Саратов--Саратовская-область"], ["Тюмень", "https://www.houzz.ru/professionals/dizayn-interyera/c/Тюмень--Тюменская-область"], ["Ижевск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Ижевск--Удмуртия"], ["Барнаул", "https://www.houzz.ru/professionals/dizayn-interyera/c/Барнаул--Алтайский-край"], ["Иркутск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Иркутск--Иркутская-область"], ["Хабаровск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Хабаровск--Хабаровский-край"], ["Ярославль", "https://www.houzz.ru/professionals/dizayn-interyera/c/Ярославль--Ярославская-область"], ["Владивосток", "https://www.houzz.ru/professionals/dizayn-interyera/c/Владивосток--Приморский-край"], ["Махачкала", "https://www.houzz.ru/professionals/dizayn-interyera/c/Махачкала--Дагестан"], ["Томск", "https://www.houzz.ru/professionals/dizayn-interyera/c/Томск--Томская-область"], ["Рязань", "https://www.houzz.ru/professionals/dizayn-interyera/c/Рязань--Рязанская-область"], ["Пенза", "https://www.houzz.ru/professionals/dizayn-interyera/c/Пенза--Пензенская-область"], ["Киров", "https://www.houzz.ru/professionals/dizayn-interyera/c/Киров--Кировская-область"], ["Чебоксары", "https://www.houzz.ru/professionals/dizayn-interyera/c/Чебоксары--Чувашия"], ["Калининград", "https://www.houzz.ru/professionals/dizayn-interyera/c/Калининград--Калининградская-область"], ["Тула", "https://www.houzz.ru/professionals/dizayn-interyera/c/Тула--Тульская-область"], ["Сочи", "https://www.houzz.ru/professionals/dizayn-interyera/c/Сочи--Краснодарский-край"], ["Ставрополь", "https://www.houzz.ru/professionals/dizayn-interyera/c/Ставрополь--Ставрополье"], ["Тверь", "https://www.houzz.ru/professionals/dizayn-interyera/c/Тверь--Тверская-область"]]
headers = {
    'Accept-Encoding': 'gzip, deflate, sdch',
    'Accept-Language': 'en-US,en;q=0.8',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
}

def parse(res, city, arhdiz):   # res - количество страниц      city:   0 - Москва    1 - Питер      arhdiz:    0 - Архитекторы    1 - Дизайнеры
    try:
        arh_houz = []
        osn_url = ''
        for i in range(res):
            if arhdiz == 0:
                if i == 0:
                    osn_url = arh_citis[city][1]
                    URL_TEMPLATE = osn_url
                else:
                    page_number = i * 15
                    URL_TEMPLATE = osn_url + "/p/" + str(page_number)
            elif arhdiz == 1:
                if i == 0:
                    osn_url = diz_citis[city][1]
                    URL_TEMPLATE = osn_url
                    URL_TEMPLATE = osn_url
                else:
                    page_number = i * 15
                    URL_TEMPLATE = osn_url + "/p/" + str(page_number)
            print(URL_TEMPLATE)
            r = requests.get(URL_TEMPLATE, headers=headers)
            time.sleep(1)
            soup = bs(r.text, "html.parser")
            arh = soup.find_all('div', class_='pro-results')[0].find_all('li', class_='ProSearchResultsV2__StyledListItem-aeq0am-0 hz-pro-search-results__item')
            for i in arh:
                    arh_houz.append(i.find_all("a")[0]["href"])
        
        print(f"Найдено профилей с {res} страниц: "+str(len(arh_houz)))
        print("Сейчас начнется сбор html-копий")
        pages = []
        chet = 0
        for i in arh_houz:
            print(f"Сбор html-копий профилей {chet}/{str(len(arh_houz))}", end="\r")
            r = requests.get(i, headers=headers)
            pages.append([r.text, i])
            chet += 1
        print(f"Сбор html-копий профилей {chet}/{str(len(arh_houz))}")
        
        #### парсинг основных данных ####
        print("\n")
        print("Сейчас начнется сбор данных...")
        data_base = []
        chet = 0
        for i in pages:
            soup = bs(i[0], "html.parser")
            #try:
            main_block = soup.find_all('div', class_='sc-183mtny-0 sc-1wm9uar-0 lnQdrs bHBpBq hui-grid')[0].find_all('p')
            main_blockk = soup.find_all('div', class_='sc-183mtny-0 sc-1wm9uar-0 lnQdrs bHBpBq hui-grid')[0]
            name = '-'
            name = main_block[0].text # name 
            
            #<h3 font-weight="bold" font-size="16px" class="sc-1hirnpv-0 jCPXIK sc-mwxddt-0 iLgjJk">Website</h3>
            website = '-'
            for c in main_blockk:
                if c.find_all("h3")[0].text == "Website":
                    website =  c.find_all("a")[0]["href"]# site
            try:
                r = requests.get(website, headers=headers)
                if not r.status_code == 200:
                        website = "-"
            except:
                try:
                    r = requests.get("https://" + website, headers=headers)
                    if not r.status_code == 200:
                        website = "-"
                        #<div class="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 hTSLyQ hNCBYN jlkdvL hui-cell">
                except:
                    website = "-"
            '''
            main_block = soup.find_all('div', class_='sc-183mtny-0 sc-1wm9uar-0 lnQdrs bHBpBq hui-grid')[0].find_all('div', class_="sc-183mtny-0 sc-1uw6j8i-0 BusinessDetails__StyledCell-sc-1iscszt-0 hTSLyQ hNCBYN jlkdvL hui-cell")
            if main_block[len(main_block)-1].find("h3").text == "Socials":
                soc = main_block[len(main_block)-1].find("a")["href"]
            else:
                soc = "-"
        '''
            data_base.append([name,  website, i[1]]) # имя, сайт, url
            chet += 1
            #except:
                #continue
            
            print(f"Парсинг основных данных {chet}/{str(len(pages))}", end="\r")
        #print(data_base)
        print(f"Парсинг основных данных {chet}/{str(len(pages))}")
        print(f"Процесс получения номеров и почты...")
        for i in range(len(data_base)):
            print(data_base[i][1])
            try:
                r = requests.get(data_base[i][1], headers=headers)
            except:
                continue
            soup = bs(r.text, "html.parser")
            kcb = []
            for b in soup.find_all('a'): # номера
                try:
                    if b["href"][0:3] == "tel":
                        #print(b["href"][3:])
                        kcb.append(b["href"][3:])
                except:
                    continue
            a = ''
            if kcb == []:
                a = "-"
            else:
                for b in kcb:
                    a += b + "\n"
            kcb = []
            data_base[i].append(a)
            for b in soup.find_all('a'): # маил
                try:
                    if b["href"][0:6] == "mailto":
                        #print(b["href"][6:]) 
                        kcb.append(b["href"][6:].replace(":", "").replace("%20", ""))
                except:
                    continue
            a = ''
            if kcb == []:
                a = "-"
            else:
                for b in kcb:
                    a += b + "\n"
            data_base[i].append(a)
            kcb = []

        #<a href="mailto:s.popova.studio@gmail.com" style="color: rgb(255, 255, 255); font-weight: 500;">s.popova.studio@gmail.com</a>
        #<a class="landing-block-node-linkcontact-link g-text-decoration-none--hover" href="tel:+79036315345" target="_blank">

        print("Last Process")
        print("Формируем Excel-таблицу")
        #имя, сайт, url
        
        wb = openpyxl.Workbook()
        list1 = wb.active
        list1.cell(row=1, column=1, value="Название")
        list1.cell(row=1, column=2, value="Сайт")
        list1.cell(row=1, column=3, value="URL")
        list1.cell(row=1, column=4, value="Номер")
        list1.cell(row=1, column=5, value="Эл. Почта")
        for i in range(len(data_base)):
            try:    
                if i == 0:
                    continue
                list1.cell(row=i+1, column=1, value=data_base[i][0])
                list1.cell(row=i+1, column=2, value=data_base[i][1])
                list1.cell(row=i+1, column=3, value=data_base[i][2])
                list1.cell(row=i+1, column=4, value=data_base[i][3])# соц сети
                list1.cell(row=i+1, column=5, value=data_base[i][4])
            except:
                continue
        name = str(random.randint(1000000, 9999999))
        wb.save(f"db_{name}_houzz.xlsx")
        print(f"Ваша база данных успешно собрана! \nНазвание: db_{name}_houzz.xlsx")
    except Exception as e:
        print(e)

print("Приветствую, это парсер сайта www.houzz.ru")
page_cite = int(input("Введите количество страниц (целое число, 1 = +- 14 страниц. Например 4 значит 56 профилей): "))
print("\n                                                                                              ", end="\r")
arh_cite = int(input("Направление (0 - Архитекторы   1 - Дизайнеры): "))
while not arh_cite == 1 and not arh_cite == 0:
    print("\n                                                                                 ", end="\r")
    arh_cite = int(input("Направление (0 - Архитекторы   1 - Дизайнеры): "))
if arh_cite == 0:
    for i in range(len(arh_citis)):
        print(f"{i}. {arh_citis[i][0]}")
elif arh_cite == 1:
    for i in range(len(diz_citis)):
        print(f"{i}. {diz_citis[i][0]}")
else:
    print("Вы выбрали неверное значение :(")
    exit()
city_cite = int(input("Введите индекс города: "))

print("\n                                                                                              ", end="\r")

print("Начинаем парсинг :D")
parse(page_cite, city_cite, arh_cite)
print("\nНажмите Enter, чтобы выйти")
input()